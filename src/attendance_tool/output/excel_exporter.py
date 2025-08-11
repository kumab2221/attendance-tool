"""Excel出力機能 - Green Phase 最小実装"""

import logging
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

from ..calculation.department_summary import DepartmentSummary
from ..calculation.summary import AttendanceSummary
from ..utils.config import ConfigManager
from .models import ConditionalFormat, ExcelExportConfig, ExportResult

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel形式でのレポート出力機能"""

    def __init__(self):
        """ExcelExporter初期化"""
        self.config_manager = ConfigManager()
        self._load_excel_config()

    def _load_excel_config(self) -> None:
        """Excel設定の読み込み"""
        try:
            # 基本設定（最小実装）
            self.excel_config = ExcelExportConfig(
                filename_pattern="attendance_report_{year}_{month:02d}.xlsx",
                worksheet_names={
                    "employee": "社員別レポート",
                    "department": "部門別レポート",
                    "summary": "サマリー",
                },
                header_style={
                    "font": {"bold": True},
                    "fill": {"patternType": "solid", "fgColor": "E6F3FF"},
                },
            )
        except Exception as e:
            logger.warning(
                f"Excel設定の読み込みに失敗しました。デフォルト設定を使用します: {e}"
            )
            self._use_default_config()

    def _use_default_config(self) -> None:
        """デフォルト設定を使用"""
        self.excel_config = ExcelExportConfig(
            filename_pattern="attendance_report_{year}_{month:02d}.xlsx",
            worksheet_names={
                "employee": "社員別レポート",
                "department": "部門別レポート",
                "summary": "サマリー",
            },
        )

    def export_excel_report(
        self,
        employee_summaries: List[AttendanceSummary],
        department_summaries: List[DepartmentSummary],
        output_path: Path,
        year: int,
        month: int,
        include_charts: bool = False,
    ) -> ExportResult:
        """Excel形式でのレポート出力"""
        start_time = time.time()

        try:
            # 出力ディレクトリの作成
            output_path = Path(output_path)
            output_path.mkdir(parents=True, exist_ok=True)

            # ファイル名生成
            filename = self.excel_config.get_filename(year, month)
            file_path = output_path / filename

            # Excelワークブック作成
            workbook = Workbook()

            # デフォルトシートを削除
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            # ワークシート作成
            self.export_employee_worksheet(workbook, employee_summaries, year, month)
            self.export_department_worksheet(
                workbook, department_summaries, year, month
            )
            self.export_summary_worksheet(
                workbook,
                employee_summaries,
                department_summaries,
                year,
                month,
                include_charts,
            )

            # ファイル保存
            workbook.save(file_path)

            # ファイルサイズ取得
            file_size = file_path.stat().st_size if file_path.exists() else 0
            processing_time = time.time() - start_time

            result = ExportResult(
                success=True,
                file_path=file_path,
                record_count=len(employee_summaries),
                file_size=file_size,
                processing_time=processing_time,
            )

            # 空データの警告
            if not employee_summaries and not department_summaries:
                result.add_warning(
                    "データが空のためヘッダーのみのファイルを作成しました"
                )

            logger.info(
                f"Excelレポートを出力しました: {file_path} ({len(employee_summaries)}件)"
            )
            return result

        except Exception as e:
            return self._handle_export_error(
                e,
                output_path / self.excel_config.get_filename(year, month),
                len(employee_summaries),
                "Excelレポート出力",
            )

    def export_employee_worksheet(
        self,
        workbook: OpenpyxlWorkbook,
        summaries: List[AttendanceSummary],
        year: int,
        month: int,
    ) -> None:
        """社員別ワークシート出力"""
        worksheet = workbook.create_sheet(self.excel_config.worksheet_names["employee"])

        # ヘッダー行定義
        headers = [
            "社員ID",
            "氏名",
            "部署",
            "対象年月",
            "出勤日数",
            "欠勤日数",
            "遅刻回数",
            "早退回数",
            "総労働時間",
            "所定労働時間",
            "残業時間",
            "深夜労働時間",
            "有給取得日数",
        ]

        # ヘッダー行書き込み
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # データ行書き込み
        for row, summary in enumerate(summaries, 2):
            row_data = self._convert_employee_summary_to_row(summary, year, month)
            for col, (header, value) in enumerate(zip(headers, row_data.values()), 1):
                worksheet.cell(row=row, column=col, value=value)

        # Excel固有機能の適用
        self._apply_excel_features(worksheet)

    def export_department_worksheet(
        self,
        workbook: OpenpyxlWorkbook,
        summaries: List[DepartmentSummary],
        year: int,
        month: int,
    ) -> None:
        """部門別ワークシート出力"""
        worksheet = workbook.create_sheet(
            self.excel_config.worksheet_names["department"]
        )

        # ヘッダー行定義
        headers = [
            "部署",
            "対象年月",
            "所属人数",
            "総出勤日数",
            "総欠勤日数",
            "総労働時間",
            "総残業時間",
            "平均出勤率",
        ]

        # ヘッダー行書き込み
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # データ行書き込み
        for row, summary in enumerate(summaries, 2):
            row_data = self._convert_department_summary_to_row(summary, year, month)
            for col, (header, value) in enumerate(zip(headers, row_data.values()), 1):
                worksheet.cell(row=row, column=col, value=value)

        # 条件付き書式の適用
        self._apply_conditional_formatting(worksheet, summaries)

        # Excel固有機能の適用
        self._apply_excel_features(worksheet)

    def export_summary_worksheet(
        self,
        workbook: OpenpyxlWorkbook,
        employee_summaries: List[AttendanceSummary],
        department_summaries: List[DepartmentSummary],
        year: int,
        month: int,
        include_charts: bool = False,
    ) -> None:
        """サマリーワークシート出力"""
        worksheet = workbook.create_sheet(self.excel_config.worksheet_names["summary"])

        # サマリー情報の計算
        total_employees = len(employee_summaries)
        total_work_days = (
            sum(s.attendance_days for s in employee_summaries)
            if employee_summaries
            else 0
        )
        avg_attendance_rate = (
            sum(d.attendance_rate for d in department_summaries)
            / len(department_summaries)
            if department_summaries
            else 0
        )
        total_overtime_hours = (
            sum(
                (s.scheduled_overtime_minutes + s.legal_overtime_minutes) / 60.0
                for s in employee_summaries
            )
            if employee_summaries
            else 0
        )
        department_count = len(department_summaries)

        # サマリー情報の書き込み
        summary_data = [
            ("総従業員数", total_employees),
            ("総出勤日数", total_work_days),
            ("平均出勤率", f"{avg_attendance_rate:.1f}%"),
            ("総残業時間", f"{total_overtime_hours:.1f}時間"),
            ("部門数", department_count),
        ]

        for row, (label, value) in enumerate(summary_data, 1):
            worksheet.cell(row=row, column=1, value=label)
            worksheet.cell(row=row, column=2, value=value)

        # グラフ作成（基本実装）
        if include_charts and department_summaries:
            self._create_department_chart(worksheet, department_summaries)

    def _apply_header_style(self, cell) -> None:
        """ヘッダーセルのスタイル適用"""
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def _apply_excel_features(self, worksheet) -> None:
        """Excel固有機能の適用"""
        # 自動フィルター設定
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = (
                f"A1:{chr(64 + worksheet.max_column)}{worksheet.max_row}"
            )

        # ウィンドウ枠固定（ヘッダー行）
        worksheet.freeze_panes = "A2"

        # 印刷設定
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        # 自動幅調整（基本実装）
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大幅制限
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _apply_conditional_formatting(
        self, worksheet, summaries: List[DepartmentSummary]
    ) -> None:
        """条件付き書式の適用"""
        if not summaries:
            return

        # 平均出勤率列（8列目）に条件付き書式を適用
        attendance_rate_column = "H"
        data_range = (
            f"{attendance_rate_column}2:{attendance_rate_column}{len(summaries) + 1}"
        )

        # 95%以上: 緑色
        green_rule = CellIsRule(
            operator="greaterThanOrEqual",
            formula=["95"],
            fill=PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            ),
        )

        # 90-95%: 黄色
        yellow_rule = CellIsRule(
            operator="between",
            formula=["90", "95"],
            fill=PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            ),
        )

        # 90%未満: 赤色
        red_rule = CellIsRule(
            operator="lessThan",
            formula=["90"],
            fill=PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            ),
        )

        worksheet.conditional_formatting.add(data_range, green_rule)
        worksheet.conditional_formatting.add(data_range, yellow_rule)
        worksheet.conditional_formatting.add(data_range, red_rule)

    def _create_department_chart(
        self, worksheet, department_summaries: List[DepartmentSummary]
    ) -> None:
        """部門別出勤率グラフの作成"""
        # グラフ用データを配置（H列以降に配置）
        chart_start_col = 10  # J列

        # ヘッダー
        worksheet.cell(row=1, column=chart_start_col, value="部門名")
        worksheet.cell(row=1, column=chart_start_col + 1, value="出勤率")

        # データ
        for row, dept in enumerate(department_summaries, 2):
            worksheet.cell(row=row, column=chart_start_col, value=dept.department_name)
            worksheet.cell(
                row=row, column=chart_start_col + 1, value=dept.attendance_rate
            )

        # 棒グラフ作成
        chart = BarChart()
        chart.title = "部門別出勤率"
        chart.x_axis.title = "部門"
        chart.y_axis.title = "出勤率(%)"

        # データ範囲設定
        data = Reference(
            worksheet,
            min_col=chart_start_col + 1,
            min_row=1,
            max_row=len(department_summaries) + 1,
        )
        categories = Reference(
            worksheet,
            min_col=chart_start_col,
            min_row=2,
            max_row=len(department_summaries) + 1,
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # グラフをワークシートに追加
        worksheet.add_chart(chart, "L2")

    def _safe_get_value(self, value: Any, default: Any) -> Any:
        """安全な値の取得（CSVExporter から共通化）"""
        if value is None or (isinstance(value, str) and not value.strip()):
            return default
        return value

    def _minutes_to_hours(self, minutes: int) -> float:
        """分を時間に換算（CSVExporter から共通化）"""
        return minutes / 60.0

    def _format_period_string(self, year: int, month: int) -> str:
        """期間文字列をフォーマット（CSVExporter から共通化）"""
        return f"{year}-{month:02d}"

    def _handle_export_error(
        self, error: Exception, file_path: Path, record_count: int, operation: str
    ) -> ExportResult:
        """Excel出力エラーの統一処理（CSVExporter から共通化）"""
        if isinstance(error, PermissionError):
            logger.error(f"{operation}でファイル書き込み権限エラー: {error}")
            error_msg = f"Permission denied: {str(error)}"
        elif isinstance(error, OSError):
            logger.error(f"{operation}でファイルシステムエラー: {error}")
            error_msg = f"OS Error: {str(error)}"
        else:
            logger.error(f"{operation}中に予期しないエラー: {error}")
            error_msg = f"Unexpected error: {str(error)}"

        result = ExportResult(
            success=False, file_path=file_path, record_count=record_count
        )
        result.add_error(error_msg)
        return result

    def _convert_employee_summary_to_row(
        self, summary: AttendanceSummary, year: int, month: int
    ) -> Dict[str, Any]:
        """AttendanceSummaryをExcel行データに変換（CSVExporter から共通化）"""
        # データバリデーション
        employee_id = self._safe_get_value(summary.employee_id, "UNKNOWN")
        employee_name = self._safe_get_value(summary.employee_name, "Unknown User")
        department = self._safe_get_value(summary.department, "未設定")

        # 欠勤日数計算
        absence_days = max(0, summary.business_days - summary.attendance_days)

        # 時間換算
        total_work_hours = self._minutes_to_hours(summary.total_work_minutes)
        overtime_hours = self._minutes_to_hours(
            summary.scheduled_overtime_minutes + summary.legal_overtime_minutes
        )

        # 標準労働時間（仮定：8時間/日）
        standard_work_hours = summary.attendance_days * 8.0

        return {
            "社員ID": employee_id,
            "氏名": employee_name,
            "部署": department,
            "対象年月": self._format_period_string(year, month),
            "出勤日数": summary.attendance_days,
            "欠勤日数": absence_days,
            "遅刻回数": summary.tardiness_count,
            "早退回数": summary.early_leave_count,
            "総労働時間": round(total_work_hours, 2),
            "所定労働時間": round(standard_work_hours, 2),
            "残業時間": round(overtime_hours, 2),
            "深夜労働時間": round(
                getattr(summary, "late_night_work_minutes", 0) / 60.0, 2
            ),
            "有給取得日数": round(summary.paid_leave_days, 1),
        }

    def _convert_department_summary_to_row(
        self, summary: DepartmentSummary, year: int, month: int
    ) -> Dict[str, Any]:
        """DepartmentSummaryをExcel行データに変換（CSVExporter から共通化）"""
        department_name = self._safe_get_value(summary.department_name, "未設定部門")

        # 時間換算
        total_work_hours = self._minutes_to_hours(summary.total_work_minutes)
        total_overtime_hours = self._minutes_to_hours(summary.total_overtime_minutes)

        # 推定値計算
        total_work_days = (
            int(summary.employee_count * summary.average_work_minutes / 480)
            if summary.average_work_minutes > 0
            else 0
        )
        total_absent_days = max(0, summary.employee_count * 22 - total_work_days)

        return {
            "部署": department_name,
            "対象年月": self._format_period_string(year, month),
            "所属人数": summary.employee_count,
            "総出勤日数": total_work_days,
            "総欠勤日数": total_absent_days,
            "総労働時間": round(total_work_hours, 2),
            "総残業時間": round(total_overtime_hours, 2),
            "平均出勤率": round(summary.attendance_rate, 1),
        }
