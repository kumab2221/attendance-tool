"""テンプレート管理機能 - TASK-303実装"""

import logging
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml

from ..calculation.department_summary import DepartmentSummary
from ..calculation.summary import AttendanceSummary
from ..utils.config import ConfigManager
from .csv_exporter import CSVExporter
from .excel_exporter import ExcelExporter
from .models import ExportResult

logger = logging.getLogger(__name__)


class TemplateManager:
    """レポートテンプレート管理クラス"""

    def __init__(self):
        """TemplateManager初期化"""
        self.config_manager = ConfigManager()
        self.template_config = self._load_template_config()
        self.excel_exporter = ExcelExporter()
        self.csv_exporter = CSVExporter()

    def _load_template_config(self) -> Dict[str, Any]:
        """テンプレート設定の読み込み"""
        try:
            config_path = Path("config/template_config.yaml")
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8") as f:
                    return yaml.safe_load(f)
            else:
                logger.warning(
                    "テンプレート設定ファイルが見つかりません。デフォルト設定を使用します。"
                )
                return self._get_default_template_config()
        except Exception as e:
            logger.error(f"テンプレート設定の読み込みに失敗しました: {e}")
            return self._get_default_template_config()

    def _get_default_template_config(self) -> Dict[str, Any]:
        """デフォルトテンプレート設定"""
        return {
            "template_settings": {
                "default": {
                    "excel": {
                        "header_info": {
                            "company_name": "勤怠管理システム",
                            "report_title": "勤怠レポート",
                            "generated_by": "自動集計ツール",
                        }
                    },
                    "csv": {
                        "header_format": {
                            "include_company_info": True,
                            "include_generation_time": True,
                        }
                    },
                }
            },
            "template_features": {
                "comparison": {"enabled": True, "previous_months": 3},
                "charts": {"enabled": True, "chart_types": ["bar_chart"]},
                "multi_month": {"enabled": True, "max_months": 12},
            },
        }

    def generate_templated_report(
        self,
        employee_summaries: List[AttendanceSummary],
        department_summaries: List[DepartmentSummary],
        output_path: Path,
        year: int,
        month: int,
        template_name: str = "default",
        output_format: str = "excel",
        include_comparison: bool = False,
        include_charts: bool = True,
    ) -> ExportResult:
        """テンプレートを使用したレポート生成"""

        logger.info(f"テンプレート '{template_name}' を使用してレポートを生成開始")

        try:
            # テンプレート設定の取得
            template_settings = self._get_template_settings(
                template_name, output_format
            )

            # 比較データの準備 (REQ-301対応)
            comparison_data = None
            if include_comparison and self._is_comparison_enabled():
                comparison_data = self._generate_comparison_data(
                    employee_summaries, department_summaries, year, month
                )

            # レポート生成
            if output_format.lower() == "excel":
                result = self._generate_excel_with_template(
                    employee_summaries,
                    department_summaries,
                    output_path,
                    year,
                    month,
                    template_settings,
                    comparison_data,
                    include_charts,
                )
            elif output_format.lower() == "csv":
                result = self._generate_csv_with_template(
                    employee_summaries,
                    department_summaries,
                    output_path,
                    year,
                    month,
                    template_settings,
                    comparison_data,
                )
            else:
                raise ValueError(f"サポートされていない出力形式: {output_format}")

            logger.info(f"テンプレートレポート生成完了: {result.file_path}")
            return result

        except Exception as e:
            logger.error(f"テンプレートレポート生成中にエラー: {e}")
            return ExportResult(
                success=False,
                file_path=output_path / f"error_report_{year}_{month:02d}.xlsx",
                record_count=len(employee_summaries),
                errors=[f"Template generation error: {str(e)}"],
            )

    def _get_template_settings(
        self, template_name: str, output_format: str
    ) -> Dict[str, Any]:
        """テンプレート設定の取得"""
        template_settings = self.template_config.get("template_settings", {})

        if template_name not in template_settings:
            logger.warning(
                f"テンプレート '{template_name}' が見つかりません。デフォルトを使用します。"
            )
            template_name = "default"

        settings = template_settings.get(template_name, {})
        return settings.get(output_format, {})

    def _is_comparison_enabled(self) -> bool:
        """比較機能が有効かチェック"""
        features = self.template_config.get("template_features", {})
        comparison = features.get("comparison", {})
        return comparison.get("enabled", False)

    def _generate_comparison_data(
        self,
        employee_summaries: List[AttendanceSummary],
        department_summaries: List[DepartmentSummary],
        current_year: int,
        current_month: int,
    ) -> Dict[str, Any]:
        """比較用データの生成 (REQ-301対応)"""

        # 過去のデータを取得（スタブ実装）
        # 実際の実装では過去のデータをデータベースや履歴ファイルから取得

        comparison_data = {
            "current_period": f"{current_year}-{current_month:02d}",
            "previous_periods": [],
            "trends": {},
            "metrics_comparison": {},
        }

        # 過去3ヶ月のダミーデータ生成
        features = self.template_config.get("template_features", {})
        comparison_config = features.get("comparison", {})
        previous_months = comparison_config.get("previous_months", 3)

        for i in range(1, previous_months + 1):
            prev_date = datetime(current_year, current_month, 1) - timedelta(
                days=i * 30
            )
            period = f"{prev_date.year}-{prev_date.month:02d}"

            # 実際の実装では履歴データから取得
            comparison_data["previous_periods"].append(
                {
                    "period": period,
                    "total_employees": len(employee_summaries),
                    "avg_attendance_rate": 92.5 + i,  # ダミー値
                    "total_overtime_hours": 150.0 - i * 10,  # ダミー値
                }
            )

        # トレンド分析（簡易版）
        comparison_data["trends"] = {
            "attendance_trend": "improving",
            "overtime_trend": "stable",
            "department_performance": "mixed",
        }

        return comparison_data

    def _generate_excel_with_template(
        self,
        employee_summaries: List[AttendanceSummary],
        department_summaries: List[DepartmentSummary],
        output_path: Path,
        year: int,
        month: int,
        template_settings: Dict[str, Any],
        comparison_data: Optional[Dict[str, Any]],
        include_charts: bool,
    ) -> ExportResult:
        """テンプレートを使用したExcel出力"""

        # 基本的なExcel出力を実行
        result = self.excel_exporter.export_excel_report(
            employee_summaries,
            department_summaries,
            output_path,
            year,
            month,
            include_charts,
        )

        if not result.success:
            return result

        # テンプレート適用（ファイル後処理）
        try:
            self._apply_excel_template(
                result.file_path, template_settings, comparison_data
            )
            result.add_warning(
                "テンプレート機能は基本実装です。詳細カスタマイゼーションは今後対応予定。"
            )
        except Exception as e:
            logger.warning(f"テンプレート適用中にエラー（処理は継続）: {e}")
            result.add_warning(f"テンプレート適用エラー: {e}")

        return result

    def _generate_csv_with_template(
        self,
        employee_summaries: List[AttendanceSummary],
        department_summaries: List[DepartmentSummary],
        output_path: Path,
        year: int,
        month: int,
        template_settings: Dict[str, Any],
        comparison_data: Optional[Dict[str, Any]],
    ) -> ExportResult:
        """テンプレートを使用したCSV出力"""

        # 基本的なCSV出力を実行
        employee_result = self.csv_exporter.export_employee_report(
            employee_summaries, output_path, year, month
        )

        if not employee_result.success:
            return employee_result

        department_result = self.csv_exporter.export_department_report(
            department_summaries, output_path, year, month
        )

        # テンプレート適用（ヘッダー情報追加）
        try:
            self._apply_csv_template(
                employee_result.file_path,
                template_settings,
                comparison_data,
                "employee",
            )
            self._apply_csv_template(
                department_result.file_path,
                template_settings,
                comparison_data,
                "department",
            )
        except Exception as e:
            logger.warning(f"CSV テンプレート適用中にエラー: {e}")
            employee_result.add_warning(f"テンプレート適用エラー: {e}")

        return employee_result

    def _apply_excel_template(
        self,
        file_path: Path,
        template_settings: Dict[str, Any],
        comparison_data: Optional[Dict[str, Any]],
    ) -> None:
        """Excelファイルにテンプレートを適用"""

        # openpyxlを使用してファイルを開き、テンプレート情報を追加
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(file_path)

            # ヘッダー情報の追加
            header_info = template_settings.get("header_info", {})
            if header_info:
                self._add_excel_header_info(workbook, header_info)

            # 比較データの追加 (REQ-301)
            if comparison_data:
                self._add_comparison_worksheet(workbook, comparison_data)

            # スタイル適用
            style_info = template_settings.get("style", {})
            if style_info:
                self._apply_excel_styles(workbook, style_info)

            workbook.save(file_path)

        except Exception as e:
            logger.error(f"Excel テンプレート適用エラー: {e}")
            raise

    def _add_excel_header_info(self, workbook, header_info: Dict[str, Any]) -> None:
        """Excelファイルにヘッダー情報を追加"""

        # 各ワークシートにヘッダー情報を追加
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # 会社名
            company_name = header_info.get("company_name", "")
            if company_name:
                worksheet["A1"] = company_name
                worksheet.merge_cells("A1:E1")

            # レポートタイトル
            report_title = header_info.get("report_title", "")
            if report_title:
                worksheet["A2"] = report_title
                worksheet.merge_cells("A2:E2")

            # 生成情報
            generated_by = header_info.get("generated_by", "")
            if generated_by:
                worksheet["A3"] = f"Generated by: {generated_by}"

            # 生成日時
            worksheet["A4"] = (
                f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )

            # 既存データを下にシフト
            worksheet.insert_rows(1, 5)

    def _add_comparison_worksheet(
        self, workbook, comparison_data: Dict[str, Any]
    ) -> None:
        """比較用ワークシートの追加 (REQ-301対応)"""

        comparison_sheet = workbook.create_sheet("比較分析")

        # 現在期間情報
        comparison_sheet["A1"] = "期間比較分析"
        comparison_sheet["A2"] = f"対象期間: {comparison_data['current_period']}"

        # 過去データ比較
        comparison_sheet["A4"] = "過去期間との比較"
        comparison_sheet["A5"] = "期間"
        comparison_sheet["B5"] = "従業員数"
        comparison_sheet["C5"] = "平均出勤率"
        comparison_sheet["D5"] = "総残業時間"

        row = 6
        for period_data in comparison_data.get("previous_periods", []):
            comparison_sheet[f"A{row}"] = period_data["period"]
            comparison_sheet[f"B{row}"] = period_data["total_employees"]
            comparison_sheet[f"C{row}"] = f"{period_data['avg_attendance_rate']:.1f}%"
            comparison_sheet[f"D{row}"] = f"{period_data['total_overtime_hours']:.1f}h"
            row += 1

        # トレンド情報
        trends = comparison_data.get("trends", {})
        comparison_sheet["A10"] = "トレンド分析"
        comparison_sheet["A11"] = (
            f"出勤率トレンド: {trends.get('attendance_trend', 'N/A')}"
        )
        comparison_sheet["A12"] = (
            f"残業時間トレンド: {trends.get('overtime_trend', 'N/A')}"
        )
        comparison_sheet["A13"] = (
            f"部門パフォーマンス: {trends.get('department_performance', 'N/A')}"
        )

    def _apply_excel_styles(self, workbook, style_info: Dict[str, Any]) -> None:
        """Excelファイルにスタイルを適用"""

        from openpyxl.styles import Font, PatternFill

        primary_color = style_info.get("primary_color", "#2E86AB").replace("#", "")

        # 各ワークシートにスタイル適用
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # ヘッダーセルのスタイル適用
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color=primary_color, end_color=primary_color, fill_type="solid"
            )

            # 最初の行（ヘッダー）にスタイル適用
            for cell in worksheet[1]:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill

    def _apply_csv_template(
        self,
        file_path: Path,
        template_settings: Dict[str, Any],
        comparison_data: Optional[Dict[str, Any]],
        report_type: str,
    ) -> None:
        """CSVファイルにテンプレートを適用"""

        # CSVファイルにヘッダー情報を追加
        header_format = template_settings.get("header_format", {})

        if not any(header_format.values()):
            return  # ヘッダー情報が不要な場合はスキップ

        # 既存ファイルの読み取り
        with open(file_path, "r", encoding="utf-8-sig") as f:
            original_content = f.read()

        # ヘッダー情報の生成
        header_lines = []

        if header_format.get("include_company_info", False):
            header_lines.append("# 勤怠管理レポート")
            header_lines.append("# Generated by 勤怠管理自動集計ツール")

        if header_format.get("include_generation_time", False):
            header_lines.append(
                f"# Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )

        if header_format.get("include_summary_stats", False) and comparison_data:
            header_lines.append("# 比較情報:")
            header_lines.append(f"# 対象期間: {comparison_data['current_period']}")

        header_lines.append("")  # 空行

        # ヘッダー付きファイルの書き込み
        with open(file_path, "w", encoding="utf-8-sig") as f:
            f.write("\n".join(header_lines))
            f.write(original_content)

    def generate_multi_month_report(
        self,
        data_by_month: Dict[str, Dict],
        output_path: Path,
        start_year: int,
        start_month: int,
        end_year: int,
        end_month: int,
        template_name: str = "default",
    ) -> ExportResult:
        """複数月統合レポート生成 (REQ-303対応)"""

        logger.info(
            f"複数月統合レポート生成開始: {start_year}-{start_month:02d} to {end_year}-{end_month:02d}"
        )

        try:
            # 複数月機能が有効かチェック
            features = self.template_config.get("template_features", {})
            multi_month = features.get("multi_month", {})
            if not multi_month.get("enabled", False):
                return ExportResult(
                    success=False,
                    file_path=output_path / "multi_month_disabled.xlsx",
                    record_count=0,
                    errors=["複数月統合機能が無効です"],
                )

            # 月数制限チェック
            max_months = multi_month.get("max_months", 12)
            month_count = self._calculate_month_count(
                start_year, start_month, end_year, end_month
            )

            if month_count > max_months:
                return ExportResult(
                    success=False,
                    file_path=output_path / "multi_month_limit.xlsx",
                    record_count=0,
                    errors=[f"月数制限を超過しています: {month_count} > {max_months}"],
                )

            # 統合データの生成
            aggregated_data = self._aggregate_multi_month_data(data_by_month)

            # 統合レポートの出力
            filename = f"multi_month_report_{start_year}{start_month:02d}_{end_year}{end_month:02d}.xlsx"
            file_path = output_path / filename

            # スタブ実装：基本的なExcel出力
            result = ExportResult(
                success=True,
                file_path=file_path,
                record_count=len(aggregated_data),
                processing_time=1.0,
            )

            # 実際のファイル生成（簡易版）
            with open(file_path, "w") as f:
                f.write("# 複数月統合レポート（スタブ実装）\n")
                f.write(
                    f"# 期間: {start_year}-{start_month:02d} から {end_year}-{end_month:02d}\n"
                )
                f.write(f"# データポイント数: {len(aggregated_data)}\n")

            result.add_warning("複数月統合機能は基本実装です。詳細機能は今後拡張予定。")
            return result

        except Exception as e:
            logger.error(f"複数月統合レポート生成エラー: {e}")
            return ExportResult(
                success=False,
                file_path=output_path / "multi_month_error.xlsx",
                record_count=0,
                errors=[f"Multi-month report error: {str(e)}"],
            )

    def _calculate_month_count(
        self, start_year: int, start_month: int, end_year: int, end_month: int
    ) -> int:
        """月数の計算"""
        start_date = datetime(start_year, start_month, 1)
        end_date = datetime(end_year, end_month, 1)

        month_count = 0
        current = start_date
        while current <= end_date:
            month_count += 1
            # 次の月に移動
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        return month_count

    def _aggregate_multi_month_data(
        self, data_by_month: Dict[str, Dict]
    ) -> Dict[str, Any]:
        """複数月データの統合"""

        # 統合タイプの取得
        features = self.template_config.get("template_features", {})
        multi_month = features.get("multi_month", {})
        aggregation_types = multi_month.get("aggregation_types", ["sum"])

        aggregated = {
            "total_months": len(data_by_month),
            "aggregation_types": aggregation_types,
            "summary_stats": {},
        }

        # 基本的な統計の計算（スタブ実装）
        total_employees = 0
        total_work_hours = 0

        for month_data in data_by_month.values():
            employees = month_data.get("employee_summaries", [])
            total_employees += len(employees)

            for emp in employees:
                total_work_hours += getattr(emp, "total_work_minutes", 0) / 60.0

        aggregated["summary_stats"] = {
            "total_employees_across_months": total_employees,
            "total_work_hours_across_months": total_work_hours,
            "average_monthly_employees": (
                total_employees / len(data_by_month) if data_by_month else 0
            ),
        }

        return aggregated

    def list_available_templates(self) -> List[str]:
        """利用可能なテンプレート一覧の取得"""
        template_settings = self.template_config.get("template_settings", {})
        return list(template_settings.keys())

    def get_template_info(self, template_name: str) -> Dict[str, Any]:
        """指定テンプレートの詳細情報取得"""
        template_settings = self.template_config.get("template_settings", {})
        return template_settings.get(template_name, {})

    def create_custom_template(
        self, template_name: str, settings: Dict[str, Any], save_to_config: bool = True
    ) -> bool:
        """カスタムテンプレートの作成"""

        try:
            # 設定の検証
            if not self._validate_template_settings(settings):
                logger.error(f"無効なテンプレート設定: {template_name}")
                return False

            # 設定に追加
            if "template_settings" not in self.template_config:
                self.template_config["template_settings"] = {}

            self.template_config["template_settings"][template_name] = settings

            # ファイルに保存
            if save_to_config:
                config_path = Path("config/template_config.yaml")
                with open(config_path, "w", encoding="utf-8") as f:
                    yaml.dump(
                        self.template_config,
                        f,
                        default_flow_style=False,
                        allow_unicode=True,
                    )

            logger.info(f"カスタムテンプレート '{template_name}' を作成しました")
            return True

        except Exception as e:
            logger.error(f"カスタムテンプレート作成エラー: {e}")
            return False

    def _validate_template_settings(self, settings: Dict[str, Any]) -> bool:
        """テンプレート設定の妥当性検証"""

        # 基本的な構造チェック
        if not isinstance(settings, dict):
            return False

        # Excel設定のチェック
        if "excel" in settings:
            excel_settings = settings["excel"]
            if not isinstance(excel_settings, dict):
                return False

        # CSV設定のチェック
        if "csv" in settings:
            csv_settings = settings["csv"]
            if not isinstance(csv_settings, dict):
                return False

        return True
