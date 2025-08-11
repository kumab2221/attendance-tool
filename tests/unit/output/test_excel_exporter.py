"""Excel出力機能のテスト - Red Phase"""

import pytest
import pandas as pd
from pathlib import Path
from datetime import datetime
from unittest.mock import Mock, patch
import tempfile
import shutil
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# テスト対象のインポート（この時点では存在しない）
from attendance_tool.output.excel_exporter import (
    ExcelExporter,
    ExcelExportConfig,
    ConditionalFormat,
)
from attendance_tool.output.models import ExportResult
from attendance_tool.calculation.summary import AttendanceSummary
from attendance_tool.calculation.department_summary import DepartmentSummary


class TestExcelExporter:
    """ExcelExporter単体テスト"""

    @pytest.fixture
    def temp_output_dir(self):
        """テスト用一時出力ディレクトリ"""
        temp_dir = tempfile.mkdtemp()
        yield Path(temp_dir)
        shutil.rmtree(temp_dir)

    @pytest.fixture
    def sample_employee_data(self):
        """サンプル社員データ"""
        from datetime import date

        return [
            AttendanceSummary(
                employee_id="EMP001",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                total_days=31,
                business_days=22,
                employee_name="田中太郎",
                department="営業部",
                attendance_days=22,
                tardiness_count=1,
                early_leave_count=0,
                total_work_minutes=10560,  # 176時間
                scheduled_overtime_minutes=960,  # 16時間
                legal_overtime_minutes=0,
                paid_leave_days=2,
            ),
            AttendanceSummary(
                employee_id="EMP002",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                total_days=31,
                business_days=22,
                employee_name="佐藤花子",
                department="開発部",
                attendance_days=20,
                tardiness_count=0,
                early_leave_count=1,
                total_work_minutes=9600,  # 160時間
                scheduled_overtime_minutes=480,  # 8時間
                legal_overtime_minutes=0,
                paid_leave_days=1,
            ),
        ]

    @pytest.fixture
    def sample_department_data(self):
        """サンプル部門データ"""
        from datetime import date

        return [
            DepartmentSummary(
                department_code="SALES",
                department_name="営業部",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                employee_count=10,
                total_work_minutes=105600,
                total_overtime_minutes=9600,
                attendance_rate=95.5,
                average_work_minutes=528,
                violation_count=0,
                compliance_score=95.0,
            ),
            DepartmentSummary(
                department_code="DEV",
                department_name="開発部",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                employee_count=8,
                total_work_minutes=76800,
                total_overtime_minutes=3840,
                attendance_rate=90.9,
                average_work_minutes=480,
                violation_count=1,
                compliance_score=90.0,
            ),
        ]

    def test_excel_exporter_initialization(self):
        """T302-001: ExcelExporter初期化テスト"""
        # この時点では ExcelExporter クラスが存在しないため失敗する
        exporter = ExcelExporter()

        assert exporter is not None
        assert hasattr(exporter, "config_manager")
        assert hasattr(exporter, "excel_config")

    def test_export_basic_excel_file(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-001: 基本Excel出力機能テスト"""
        exporter = ExcelExporter()

        # Excel出力実行
        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        # 結果検証
        assert isinstance(result, ExportResult)
        assert result.success is True
        assert result.file_path.name == "attendance_report_2024_01.xlsx"
        assert result.file_path.exists()
        assert result.record_count == len(sample_employee_data)
        assert result.file_size > 0
        assert result.processing_time >= 0

    def test_employee_worksheet_structure(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-002: 社員別ワークシート構造テスト"""
        exporter = ExcelExporter()

        # Excel出力実行
        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        # Excelファイルを開いて検証
        workbook = load_workbook(result.file_path)

        # 社員別ワークシートの存在確認
        assert "社員別レポート" in workbook.sheetnames
        worksheet = workbook["社員別レポート"]

        # ヘッダー行の確認（13カラム）
        expected_headers = [
            "社員ID",
            "氏名",
            "部署",
            "対象年月",
            "出勤日数",
            "欠勤日数",
            "遅刻回数",
            "早退回数",
            "総労働時間",
            "所定労働時間",
            "残業時間",
            "深夜労働時間",
            "有給取得日数",
        ]

        for i, expected_header in enumerate(expected_headers, 1):
            assert worksheet.cell(row=1, column=i).value == expected_header

        # データ行の確認
        assert worksheet.max_row == len(sample_employee_data) + 1  # ヘッダー + データ行

        # 具体的なデータ内容確認
        assert worksheet.cell(row=2, column=1).value == "EMP001"
        assert worksheet.cell(row=2, column=2).value == "田中太郎"
        assert worksheet.cell(row=2, column=3).value == "営業部"

    def test_employee_worksheet_formatting(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-002: 社員別ワークシート書式テスト"""
        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        workbook = load_workbook(result.file_path)
        worksheet = workbook["社員別レポート"]

        # ヘッダー行の書式確認
        header_cell = worksheet.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.index is not None  # 背景色設定確認

        # 数値列の書式確認
        work_hours_cell = worksheet.cell(row=2, column=9)  # 総労働時間
        assert isinstance(work_hours_cell.value, (int, float))

        # 自動幅調整の確認（列幅が初期値より大きい）
        for column in worksheet.columns:
            col_letter = column[0].column_letter
            assert worksheet.column_dimensions[col_letter].width > 0

    def test_department_worksheet_structure(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-003: 部門別ワークシート構造テスト"""
        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        workbook = load_workbook(result.file_path)

        # 部門別ワークシートの存在確認
        assert "部門別レポート" in workbook.sheetnames
        worksheet = workbook["部門別レポート"]

        # ヘッダー行の確認（8カラム）
        expected_headers = [
            "部署",
            "対象年月",
            "所属人数",
            "総出勤日数",
            "総欠勤日数",
            "総労働時間",
            "総残業時間",
            "平均出勤率",
        ]

        for i, expected_header in enumerate(expected_headers, 1):
            assert worksheet.cell(row=1, column=i).value == expected_header

        # データ行の確認
        assert worksheet.max_row == len(sample_department_data) + 1

        # 具体的なデータ内容確認
        assert worksheet.cell(row=2, column=1).value == "営業部"
        assert worksheet.cell(row=2, column=3).value == 10  # 所属人数

    def test_department_conditional_formatting(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-003: 部門別条件付き書式テスト"""
        # 条件付き書式用のテストデータを作成
        from datetime import date

        department_data_with_various_rates = [
            DepartmentSummary(
                department_code="EXCELLENT",
                department_name="優秀部門",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                employee_count=5,
                total_work_minutes=52800,
                total_overtime_minutes=2400,
                attendance_rate=97.0,  # 95%以上 → 緑色
                average_work_minutes=528,
                violation_count=0,
                compliance_score=97.0,
            ),
            DepartmentSummary(
                department_code="AVERAGE",
                department_name="普通部門",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                employee_count=8,
                total_work_minutes=76800,
                total_overtime_minutes=3840,
                attendance_rate=92.0,  # 90-95% → 黄色
                average_work_minutes=480,
                violation_count=1,
                compliance_score=92.0,
            ),
            DepartmentSummary(
                department_code="NEEDSIMPRV",
                department_name="改善必要部門",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                employee_count=3,
                total_work_minutes=28800,
                total_overtime_minutes=0,
                attendance_rate=88.0,  # 90%未満 → 赤色
                average_work_minutes=480,
                violation_count=3,
                compliance_score=88.0,
            ),
        ]

        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=department_data_with_various_rates,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        workbook = load_workbook(result.file_path)
        worksheet = workbook["部門別レポート"]

        # 条件付き書式の適用確認
        assert len(worksheet.conditional_formatting) > 0

        # 平均出勤率列（8列目）に条件付き書式が適用されていることを確認
        attendance_rate_range = worksheet.conditional_formatting
        assert any(cf for cf in attendance_rate_range if "H" in str(cf.cells))

    def test_summary_worksheet_creation(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-004: サマリーワークシート作成テスト"""
        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        workbook = load_workbook(result.file_path)

        # サマリーワークシートの存在確認
        assert "サマリー" in workbook.sheetnames
        worksheet = workbook["サマリー"]

        # 集計情報の確認
        expected_metrics = [
            "総従業員数",
            "総出勤日数",
            "平均出勤率",
            "総残業時間",
            "部門数",
        ]

        # 各メトリクスが存在することを確認（具体的な位置は実装に依存）
        all_values = []
        for row in worksheet.iter_rows(values_only=True):
            all_values.extend([cell for cell in row if cell is not None])

        for metric in expected_metrics:
            assert any(metric in str(value) for value in all_values)

    def test_summary_worksheet_charts(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-004: サマリーワークシートグラフテスト"""
        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
            include_charts=True,
        )

        workbook = load_workbook(result.file_path)
        worksheet = workbook["サマリー"]

        # グラフオブジェクトの存在確認
        assert len(worksheet._charts) > 0

        # 部門別出勤率棒グラフの確認
        chart = worksheet._charts[0]
        assert chart.graphical_properties is not None

    def test_excel_specific_features(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-005: Excel固有機能テスト"""
        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        workbook = load_workbook(result.file_path)

        # 各ワークシートのExcel固有機能確認
        for sheet_name in ["社員別レポート", "部門別レポート"]:
            worksheet = workbook[sheet_name]

            # 自動フィルター設定確認
            assert worksheet.auto_filter is not None

            # ウィンドウ枠固定確認
            assert worksheet.freeze_panes is not None

            # 印刷設定確認
            assert worksheet.page_setup.fitToWidth == 1
            assert worksheet.page_setup.fitToHeight == 0

    def test_export_with_empty_data(self, temp_output_dir):
        """T302-E001: 空データ処理テスト"""
        exporter = ExcelExporter()

        # 空データでの出力テスト
        result = exporter.export_excel_report(
            employee_summaries=[],
            department_summaries=[],
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        # 結果検証
        assert result.success is True
        assert result.file_path.exists()
        assert len(result.warnings) > 0  # 空データの警告があること

        # ファイル内容確認
        workbook = load_workbook(result.file_path)

        # ワークシートは作成されているがデータ行はない
        employee_sheet = workbook["社員別レポート"]
        assert employee_sheet.max_row == 1  # ヘッダーのみ

        department_sheet = workbook["部門別レポート"]
        assert department_sheet.max_row == 1  # ヘッダーのみ

    def test_export_permission_error(
        self, sample_employee_data, sample_department_data
    ):
        """T302-E002: ファイル権限エラーテスト"""
        exporter = ExcelExporter()

        # 存在しない/書き込み不可のパス
        invalid_path = Path("/nonexistent/directory")

        result = exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=invalid_path,
            year=2024,
            month=1,
        )

        # エラー結果の検証
        assert result.success is False
        assert any("Permission denied" in error for error in result.errors)

    def test_export_with_invalid_data(self, temp_output_dir):
        """T302-E003: 不正データ処理テスト"""
        # 不正データを含む社員データ
        from datetime import date

        invalid_employee_data = [
            AttendanceSummary(
                employee_id=None,  # None値
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                total_days=31,
                business_days=22,
                employee_name="",  # 空文字列
                department=None,
                attendance_days=-1,  # 負の値
                tardiness_count=0,
                early_leave_count=0,
                total_work_minutes=-480,  # 負の値
                scheduled_overtime_minutes=0,
                legal_overtime_minutes=0,
                paid_leave_days=0,
            )
        ]

        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=invalid_employee_data,
            department_summaries=[],
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        # 部分的成功とデフォルト値適用の確認
        assert result.success is True
        assert len(result.warnings) > 0

        # ファイル内容確認
        workbook = load_workbook(result.file_path)
        worksheet = workbook["社員別レポート"]

        # デフォルト値が適用されていることを確認
        assert worksheet.cell(row=2, column=1).value == "UNKNOWN"  # employee_id
        assert worksheet.cell(row=2, column=2).value == "Unknown User"  # employee_name

    def test_large_data_processing(self, temp_output_dir):
        """T302-B001: 大容量データ処理テスト"""
        # 大量のデータを生成（メモリ制約のためサイズは調整）
        from datetime import date

        large_employee_data = []
        for i in range(1000):  # 1000名分
            large_employee_data.append(
                AttendanceSummary(
                    employee_id=f"EMP{i:04d}",
                    period_start=date(2024, 1, 1),
                    period_end=date(2024, 1, 31),
                    total_days=31,
                    business_days=22,
                    employee_name=f"社員{i}",
                    department=f"部門{i % 10}",
                    attendance_days=22,
                    tardiness_count=0,
                    early_leave_count=0,
                    total_work_minutes=10560,
                    scheduled_overtime_minutes=480,
                    legal_overtime_minutes=0,
                    paid_leave_days=1,
                )
            )

        exporter = ExcelExporter()

        import time

        start_time = time.time()

        result = exporter.export_excel_report(
            employee_summaries=large_employee_data,
            department_summaries=[],
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        processing_time = time.time() - start_time

        # パフォーマンス基準確認
        assert result.success is True
        assert processing_time < 60  # 60秒以内
        assert result.file_size < 20 * 1024 * 1024  # 20MB以下

    def test_unicode_character_handling(self, temp_output_dir):
        """T302-B004: 特殊文字処理テスト"""
        from datetime import date

        unicode_employee_data = [
            AttendanceSummary(
                employee_id="EMP001",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                total_days=31,
                business_days=22,
                employee_name="田中🌸太郎",  # 絵文字
                department="R&D★部",  # 特殊記号
                attendance_days=22,
                tardiness_count=0,
                early_leave_count=0,
                total_work_minutes=10560,
                scheduled_overtime_minutes=480,
                legal_overtime_minutes=0,
                paid_leave_days=1,
            ),
            AttendanceSummary(
                employee_id="EMP002",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                total_days=31,
                business_days=22,
                employee_name="Smith, John Jr.",  # 英語名
                department="総務/人事",  # スラッシュ
                attendance_days=20,
                tardiness_count=1,
                early_leave_count=0,
                total_work_minutes=9600,
                scheduled_overtime_minutes=0,
                legal_overtime_minutes=0,
                paid_leave_days=2,
            ),
        ]

        exporter = ExcelExporter()

        result = exporter.export_excel_report(
            employee_summaries=unicode_employee_data,
            department_summaries=[],
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        assert result.success is True

        # Unicode文字が正確に保存されていることを確認
        workbook = load_workbook(result.file_path)
        worksheet = workbook["社員別レポート"]

        assert worksheet.cell(row=2, column=2).value == "田中🌸太郎"
        assert worksheet.cell(row=2, column=3).value == "R&D★部"
        assert worksheet.cell(row=3, column=2).value == "Smith, John Jr."
        assert worksheet.cell(row=3, column=3).value == "総務/人事"


class TestExcelExportConfig:
    """ExcelExportConfig設定テスト"""

    def test_excel_config_initialization(self):
        """Excel設定初期化テスト"""
        config = ExcelExportConfig(
            filename_pattern="test_{year}_{month:02d}.xlsx",
            worksheet_names={
                "employee": "社員別レポート",
                "department": "部門別レポート",
                "summary": "サマリー",
            },
            header_style={
                "font": {"bold": True},
                "fill": {"patternType": "solid", "fgColor": "E6F3FF"},
            },
        )

        assert config.filename_pattern == "test_{year}_{month:02d}.xlsx"
        assert config.get_filename(2024, 1) == "test_2024_01.xlsx"
        assert config.worksheet_names["employee"] == "社員別レポート"

    def test_conditional_format_definition(self):
        """条件付き書式設定テスト"""
        conditional_format = ConditionalFormat(
            column="attendance_rate",
            condition_type="between",
            values=[90, 95],
            format_style={"fill": {"patternType": "solid", "fgColor": "FFFF00"}},
        )

        assert conditional_format.column == "attendance_rate"
        assert conditional_format.condition_type == "between"
        assert conditional_format.values == [90, 95]


class TestExcelIntegration:
    """Excel出力統合テスト"""

    def test_csv_excel_consistency(
        self, temp_output_dir, sample_employee_data, sample_department_data
    ):
        """T302-I001: CSV出力との一貫性テスト"""
        from attendance_tool.output.csv_exporter import CSVExporter

        # CSV出力
        csv_exporter = CSVExporter()
        csv_result = csv_exporter.export_employee_report(
            summaries=sample_employee_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        # Excel出力
        excel_exporter = ExcelExporter()
        excel_result = excel_exporter.export_excel_report(
            employee_summaries=sample_employee_data,
            department_summaries=sample_department_data,
            output_path=temp_output_dir,
            year=2024,
            month=1,
        )

        assert csv_result.success is True
        assert excel_result.success is True

        # データ一貫性の確認（CSVとExcelで同じデータが出力されている）
        csv_df = pd.read_csv(csv_result.file_path)

        workbook = load_workbook(excel_result.file_path)
        worksheet = workbook["社員別レポート"]

        # 社員数の一貫性確認
        assert len(csv_df) == worksheet.max_row - 1  # ヘッダー除く

        # 具体的なデータ値の一貫性確認
        assert csv_df.iloc[0]["社員ID"] == worksheet.cell(row=2, column=1).value


# パフォーマンステスト用のマーカー
pytest.mark.performance = pytest.mark.slowtest


@pytest.mark.performance
class TestExcelPerformance:
    """Excelパフォーマンステスト"""

    def test_processing_time_measurement(self, temp_output_dir):
        """T302-P001: 処理時間測定テスト"""
        # 様々なサイズのデータでの処理時間測定
        test_cases = [
            (10, 5),  # 10名, 5部門 < 5秒
            (100, 10),  # 100名, 10部門 < 10秒
        ]

        for employee_count, department_count in test_cases:
            employee_data = self._generate_test_employee_data(employee_count)
            department_data = self._generate_test_department_data(department_count)

            exporter = ExcelExporter()

            import time

            start_time = time.time()

            result = exporter.export_excel_report(
                employee_summaries=employee_data,
                department_summaries=department_data,
                output_path=temp_output_dir,
                year=2024,
                month=1,
            )

            processing_time = time.time() - start_time

            assert result.success is True

            # パフォーマンス基準
            if employee_count == 10:
                assert processing_time < 5
            elif employee_count == 100:
                assert processing_time < 10

    def _generate_test_employee_data(self, count):
        """テスト用社員データ生成"""
        from datetime import date

        return [
            AttendanceSummary(
                employee_id=f"EMP{i:04d}",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                total_days=31,
                business_days=22,
                employee_name=f"社員{i}",
                department=f"部門{i % 5}",
                attendance_days=22,
                tardiness_count=0,
                early_leave_count=0,
                total_work_minutes=10560,
                scheduled_overtime_minutes=480,
                legal_overtime_minutes=0,
                paid_leave_days=1,
            )
            for i in range(count)
        ]

    def _generate_test_department_data(self, count):
        """テスト用部門データ生成"""
        from datetime import date

        return [
            DepartmentSummary(
                department_code=f"DEPT{i:03d}",
                department_name=f"部門{i}",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 31),
                employee_count=10,
                total_work_minutes=105600,
                total_overtime_minutes=4800,
                attendance_rate=95.0,
                average_work_minutes=528,
                violation_count=0,
                compliance_score=95.0,
            )
            for i in range(count)
        ]
